#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""fill_ip.py — 把审核后的记录清单写入知识产权月度统计/累计表

用法:
  python fill_ip.py --records records.json --xlsx 目标表.xlsx [--fix fix.json] [--rebuild-monthly]

records.json 格式（AI 审核 parse_ip.py 输出后整理）:
{
  "1-国内专利-受理": [{"专利名称": "...", "专利类型": "发明", "申请人": "...",
                        "申请日期": "2024-01-26", "申请号": "202410115098.9",
                        "发明/设计人": "...", "专利权人": "..."}, ...],
  "2-国内专利-授权": [{"...": "...", "授权公告日": "2025-10-10"}, ...],
  "3-海外专利PCT阶段": [{"...": "...", "国际申请号": "PCT/CN2024/118968", "国际申请日": "..."}, ...],
  "5-软著": [{"软件著作权名称": "...", "证书编号": "...", "登记号": "...", "著作权人": "...", "登记日期": "...", ...}, ...]
}

fix.json（可选，配置区固定值，缺省用内置示例值）:
{"部门/单位": "XX服务中心", "使用保管人": "张三", "技术领域(一级)": "...", ...}

安全: 自动检测每个 sheet 已有数据行数，新记录从下一个空行写入，绝不覆盖已有行；序号自动顺延。
"""
import json, sys, argparse, os
from collections import defaultdict
import openpyxl

DEFAULT_FIX = {
    "部门/单位": "XX服务中心（示例）", "使用保管人": "张三（示例）",
    "技术领域(一级)": "电气技术", "技术领域(二级)": "供电、配电、用电与电气化",
    "技术领域(三级)": "用电管理", "技术领域(四级)": "其他",
    "智能电网环节": "智能用电", "知识产权增加方式": "科技信息项目自主开发",
    "项目类型": "005||科技项目", "上报年份": None, "上报季度": None,
}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--records", required=True, help="记录清单 JSON")
    ap.add_argument("--xlsx", required=True, help="目标 xlsx")
    ap.add_argument("--fix", help="配置区固定值 JSON（可选）")
    ap.add_argument("--rebuild-monthly", action="store_true", help="重建「月度统计」sheet")
    args = ap.parse_args()

    records = json.load(open(args.records, encoding="utf-8"))
    fix = dict(DEFAULT_FIX)
    if args.fix:
        fix.update(json.load(open(args.fix, encoding="utf-8")))

    wb = openpyxl.load_workbook(args.xlsx)

    for sheet_name, rows in records.items():
        if sheet_name not in wb.sheetnames:
            print(f"!! 跳过（sheet 不存在）: {sheet_name}")
            continue
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        idx = {h: i for i, h in enumerate(headers)}
        # 检测已有数据行数
        n = sum(1 for row in ws.iter_rows(min_row=2, values_only=True)
                if any(v is not None and str(v).strip() for v in row))
        r = n + 1  # 下一个空行的数据行号（行号从2起算数据）
        for row in rows:
            r += 1
            vals = {k: v for k, v in fix.items() if v is not None}
            vals.update(row)
            vals["序号"] = r - 1
            for k, v in vals.items():
                if k in idx:
                    ws.cell(row=r, column=idx[k] + 1, value=v)
        print(f"{sheet_name}: 已有 {n} 条，新增 {len(rows)} 条 -> 共 {n + len(rows)} 条")

    if args.rebuild_monthly:
        rebuild_monthly(wb)

    wb.save(args.xlsx)
    print("saved OK ->", args.xlsx)


def rebuild_monthly(wb):
    """重建「月度统计」：按申请/登记月汇总各类别数量"""
    if "月度统计" not in wb.sheetnames:
        ws = wb.create_sheet("月度统计")
        ws.append(["月份", "专利受理", "专利授权", "软著", "论文", "论著", "合计"])
    else:
        ws = wb["月度统计"]
    ws.delete_rows(2, ws.max_row)

    counts = defaultdict(lambda: [0, 0, 0, 0, 0])
    date_cols = {
        "1-国内专利-受理": ("申请日期", 0),
        "2-国内专利-授权": ("申请日期", 1),
        "3-海外专利PCT阶段": ("国际申请日", 0),
        "5-软著": ("登记日期", 2),
    }
    for sheet_name, (dcol, kind) in date_cols.items():
        if sheet_name not in wb.sheetnames:
            continue
        sh = wb[sheet_name]
        h = [c.value for c in sh[1]]
        if dcol not in h:
            continue
        di = h.index(dcol)
        for row in sh.iter_rows(min_row=2, values_only=True):
            if row[di] and "-" in str(row[di])[:7]:
                counts[str(row[di])[:7]][kind] += 1

    for m in sorted(counts):
        a, b, c, d, e = counts[m]
        ws.append([m, a, b, c, d, e, a + b + c + d + e])
    print("月度统计重建:", len(counts), "个月")


if __name__ == "__main__":
    main()
